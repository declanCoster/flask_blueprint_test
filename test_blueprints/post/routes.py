from flask import request, Blueprint, jsonify
import openpyxl
import logging

post = Blueprint('post', __name__)

def check_exists(sheet, city):
    for i, xl_city in enumerate(sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True)):
        if city == xl_city[0]:
            return i + 2
    return -1

@post.route('/update', methods=['POST'])
def post_update():
    logging.info('/update request made')
    # processing parameters
    params = request.get_json()
    keys = list(params.keys())
    needed_keys = ['city', 'temperature', 'altitude']
    missing_keys = [value for value in needed_keys if value not in keys]
    parameter_count_check = len(keys) + len(missing_keys)
    # parameter checking
    message = ''
    if len(missing_keys) != 0:
        message += 'the following keys are missing: ' + ' '.join(missing_keys)
    if parameter_count_check != 3:
        message += 'Unnecessary parameters were sent with the request.'
    if message != '':
        logging.debug(message)
        return (jsonify({"message": message}), 400, {'Access-Control-Allow-Origin': '*'})
    # completing request
    wb = openpyxl.load_workbook('/home/declan/flask_blueprint_test/test_blueprints/test.xlsx')
    sheet = wb.active
    index = check_exists(sheet, params['city'])
    if index < 0:       # city not already in system
        logging.info(f'new city, {params["city"]}, added.')
        new_row = sheet.max_row + 1
        sheet.cell(row=new_row, column=1, value=params['city'])
        sheet.cell(row=new_row, column=2, value=params['temperature'])
        sheet.cell(row=new_row, column=3, value=params['altitude'])
    else:
        logging.info(f'{params["city"]} updated.')
        sheet.cell(row=index, column=2, value=params['temperature'])
        sheet.cell(row=index, column=3, value=params['altitude'])
    wb.save(filename='/home/declan/flask_blueprint_test/test_blueprints/test.xlsx')
    logging.info('changes saved to excel.')
    return (jsonify({"message": 'Update complete!'}), 200, {'Access-Control-Allow-Origin': '*'})
