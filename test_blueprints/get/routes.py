from flask import request, Blueprint, jsonify
import openpyxl
import logging

get = Blueprint('get', __name__)

def get_city_info(sheet):
    rows = sheet.max_row
    city_info = {}
    attributes = [sheet.cell(row=1, column=2).value, sheet.cell(row=1, column=3).value]
    for row in range(2, rows+1):
        city_name = sheet.cell(row=row, column=1).value
        city_info[city_name] = {attributes[0]: sheet.cell(row=row, column=2).value, 
                                attributes[1]: sheet.cell(row=row, column=3).value}
    return city_info

@get.route('/all', methods=['GET'])
def get_all():
    logging.info('/all request made')
    wb = openpyxl.load_workbook('/home/declan/flask_blueprint_test/test_blueprints/test.xlsx')
    sheet = wb.active
    return (jsonify(get_city_info(sheet)), 200, {'Access-Control-Allow-Origin': '*'})


@get.route('/city', methods=['GET'])
def get_city():
    logging.info('/city request made')
    params = request.get_json()
    keys = list(params.keys())
    if 'city' not in keys:
        logging.debug(f'city not sent as a parameter')
        return (jsonify({"message": f'city not sent as a parameter'}), 400, {'Access-Control-Allow-Origin': '*'})
    elif len(keys) != 1:
        logging.debug('more than just city parameter sent')
        return (jsonify({"message": 'more than just city parameter sent'}), 400, {'Access-Control-Allow-Origin': '*'})
    else:
        wb = openpyxl.load_workbook('/home/declan/flask_blueprint_test/test_blueprints/test.xlsx')
        sheet = wb.active
        city_info = get_city_info(sheet)
        if params['city'] in city_info.keys():
            logging.info(f'city request completed for {params["city"]}')
            return (jsonify(city_info[params['city']]), 200, {'Access-Control-Allow-Origin': '*'})
        else:
            logging.debug(f"{params['city']} is not in the system")
            return (jsonify({"message": f"{params['city']} is not in the system"}), 400, {'Access-Control-Allow-Origin': '*'})

